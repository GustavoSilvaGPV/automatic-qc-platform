from fastapi import FastAPI, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from io import BytesIO
import pandas as pd
import re
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

latest_report_path = "QC_Report.xlsx"

# -----------------------------------
# NORMALIZATION
# -----------------------------------

def normalize_column_name(name):

    if pd.isna(name):
        return ""

    name = str(name).lower()

    name = name.replace(" ", "")
    name = name.replace("_", "")
    name = name.replace("-", "")

    return name


def row_to_text(row_values):

    parts = []

    for value in row_values:

        if pd.isna(value):
            continue

        parts.append(str(value))

    return " ".join(parts).lower()


def detect_header_row(raw_df):

    for idx, row in raw_df.iterrows():

        row_text = row_to_text(row.tolist())

        has_length = any(
            token in row_text
            for token in ("length", "lenght", "laenge", "lange", "geolength")
        )

        has_mrv = "mrv" in row_text

        has_duct = bool(re.search(r"\d+x\d+", row_text)) or "duct" in row_text

        if has_length and (has_mrv or has_duct):
            return idx

    return 0


# -----------------------------------
# COLUMN ALIASES
# -----------------------------------

COLUMN_ALIASES = {

    "length": [
        "length",
        "lenght",
        "geolength",
        "laenge",
        "lange"
    ],

    "mrv": [
        "mrv"
    ],

    "hausanschl": [
        "hausanschl",
        "hausanschluss",
        "hausanschluß"
    ],
}


# -----------------------------------
# FIND COLUMN
# -----------------------------------

def find_column(df_columns, target):

    aliases = COLUMN_ALIASES.get(target, [])

    for column in df_columns:

        normalized = normalize_column_name(column)

        for alias in aliases:

            if normalized == normalize_column_name(alias):
                return column

    return None


# -----------------------------------
# DETECT DUCT COLUMNS
# -----------------------------------

def detect_duct_columns(df_columns):

    duct_columns = []

    for column in df_columns:

        column_str = str(column).lower()

        if re.search(r"\d+x\d+", column_str):

            duct_columns.append(column)

        elif "duct" in column_str:

            duct_columns.append(column)

    return duct_columns


# -----------------------------------
# STYLE EXCEL
# -----------------------------------

def style_sheet(ws, color):

    header_fill = PatternFill(
        start_color=color,
        end_color=color,
        fill_type="solid"
    )

    border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000")
    )

    for cell in ws[1]:

        cell.fill = header_fill
        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )

        cell.border = border

    for row in ws.iter_rows():

        for cell in row:
            cell.border = border

    for column_cells in ws.columns:

        length = max(
            len(str(cell.value))
            if cell.value else 0
            for cell in column_cells
        )

        adjusted_width = length + 5

        ws.column_dimensions[
            get_column_letter(column_cells[0].column)
        ].width = adjusted_width


@app.post("/upload")
async def upload_file(file: UploadFile = File(...)):

    global latest_report_path

    contents = await file.read()
    buffer = BytesIO(contents)

    excel_file = pd.ExcelFile(buffer)

    sheet_names = excel_file.sheet_names

    qc_results = []

    workbook_structure = []

    # -----------------------------------
    # PROCESS SHEETS
    # -----------------------------------

    for sheet in sheet_names:

        buffer.seek(0)
        raw_df = pd.read_excel(
            buffer,
            sheet_name=sheet,
            header=None,
        )

        header_row = detect_header_row(raw_df)

        buffer.seek(0)
        df = pd.read_excel(
            buffer,
            sheet_name=sheet,
            header=header_row,
        )

        columns = df.columns.tolist()

        workbook_structure.append({
            "sheet": sheet,
            "columns": columns
        })

        duct_columns = detect_duct_columns(columns)

        length_column = find_column(
            columns,
            "length"
        )

        mrv_column = find_column(
            columns,
            "mrv"
        )

        hausanschl_column = find_column(
            columns,
            "hausanschl"
        )

        is_trench_like = (
            len(duct_columns) > 0
            and
            length_column is not None
        )

        if not is_trench_like:
            continue

        # -----------------------------------
        # VALIDATE ROWS
        # -----------------------------------

        for index, row in df.iterrows():

            try:

                duct_total = 0

                for duct_column in duct_columns:

                    value = row.get(
                        duct_column,
                        0
                    )

                    if pd.isna(value):
                        value = 0

                    try:
                        duct_total += float(value)
                    except:
                        pass

                # LENGTH

                length = 0

                if length_column:

                    length = row.get(
                        length_column,
                        0
                    )

                    if pd.isna(length):
                        length = 0

                    length = float(length)

                # MRV

                mrv = 0

                if mrv_column:

                    mrv = row.get(
                        mrv_column,
                        0
                    )

                    if pd.isna(mrv):
                        mrv = 0

                    mrv = float(mrv)

                expected_mrv = round(
                    duct_total * length,
                    2
                )

                # MRV CHECK

                if mrv_column:

                    if round(mrv, 2) == expected_mrv:

                        qc_results.append({
                            "sheet": sheet,
                            "row": index + 2,
                            "status": "PASSED",
                            "type": "MRV",
                            "message": f"MRV correct ({mrv})"
                        })

                    else:

                        qc_results.append({
                            "sheet": sheet,
                            "row": index + 2,
                            "status": "FAILED",
                            "type": "MRV",
                            "message": f"Expected {expected_mrv}, Found {mrv}"
                        })

                # HAUSANSCHL CHECK

                if hausanschl_column:

                    hausanschl = str(
                        row.get(
                            hausanschl_column,
                            ""
                        )
                    ).strip()

                    if hausanschl.lower() == "no":

                        if duct_total <= 0 or mrv <= 0:

                            qc_results.append({
                                "sheet": sheet,
                                "row": index + 2,
                                "status": "FAILED",
                                "type": "Hausanschl",
                                "message": "Hausanschl No but ducts/MRV invalid"
                            })

                        else:

                            qc_results.append({
                                "sheet": sheet,
                                "row": index + 2,
                                "status": "PASSED",
                                "type": "Hausanschl",
                                "message": "Hausanschl No correct"
                            })

                    else:

                        if duct_total == 0 and mrv == 0:

                            qc_results.append({
                                "sheet": sheet,
                                "row": index + 2,
                                "status": "PASSED",
                                "type": "Hausanschl",
                                "message": "Special connection correct"
                            })

                        else:

                            qc_results.append({
                                "sheet": sheet,
                                "row": index + 2,
                                "status": "FAILED",
                                "type": "Hausanschl",
                                "message": "Special connection invalid"
                            })

            except Exception as e:

                qc_results.append({
                    "sheet": sheet,
                    "row": index + 2,
                    "status": "FAILED",
                    "type": "System",
                    "message": str(e)
                })

    # -----------------------------------
    # EXPORT REPORT
    # -----------------------------------

    passed_results = [
        r for r in qc_results
        if r["status"] == "PASSED"
    ]

    failed_results = [
        r for r in qc_results
        if r["status"] == "FAILED"
    ]

    summary_df = pd.DataFrame([{
        "Total Checks": len(qc_results),
        "Passed": len(passed_results),
        "Failed": len(failed_results),
        "Generated": datetime.now()
    }])

    passed_df = pd.DataFrame(passed_results)

    failed_df = pd.DataFrame(failed_results)

    original_name = file.filename.replace(".xlsx", "")

    latest_report_path = (
        f"{original_name}_QC_Report.xlsx"
    )

    with pd.ExcelWriter(
        latest_report_path,
        engine="openpyxl"
    ) as writer:

        summary_df.to_excel(
            writer,
            sheet_name="Summary",
            index=False
        )

        failed_df.to_excel(
            writer,
            sheet_name="Errors",
            index=False
        )

        passed_df.to_excel(
            writer,
            sheet_name="Passed",
            index=False
        )

    wb = load_workbook(latest_report_path)

    style_sheet(
        wb["Summary"],
        "4F81BD"
    )

    style_sheet(
        wb["Errors"],
        "C0504D"
    )

    style_sheet(
        wb["Passed"],
        "9BBB59"
    )

    wb.save(latest_report_path)

    return {
        "filename": file.filename,
        "results": qc_results,
        "structure": workbook_structure,
        "summary": {
            "total": len(qc_results),
            "passed": len(passed_results),
            "failed": len(failed_results)
        }
    }


@app.get("/download-report")
async def download_report():

    return FileResponse(
        latest_report_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=latest_report_path
    )